import os
from uuid import uuid4
from tempfile import gettempdir
from typing import Union, Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from classes import Model, ModelHistory
from loader import db


class ExcelHandler:
    def __init__(self, input_file_path: str, open_model_history_file: str, input_sheet_name: str = "input", models_start_cell: tuple = (4, 1), models_end_cell: tuple = (100, 20),
                 markets_names_row: int = 3, models_names_column: int = 1, temp_directory: str = gettempdir()):
        self.file = input_file_path
        self.open_model_history_file = open_model_history_file
        self.workbook = openpyxl.load_workbook(self.file, keep_vba=True)

        self.models_start_cell = models_start_cell
        self.models_end_cell = models_end_cell

        self.markets_names_row = markets_names_row
        self.models_names_column = models_names_column

        self.input_sheet_name = input_sheet_name
        self._check_input_sheet()

        self.temp_directory = temp_directory

    def _check_input_sheet(self) -> None:
        if self.input_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Input sheet \"{self.input_sheet_name}\" not found!")

    # @property
    # def workbook(self) -> Workbook:
    #     return openpyxl.load_workbook(self.file, keep_vba=True)

    def save(self):
        self.workbook.save(self.file)

    @property
    def input_sheet(self) -> Worksheet:
        return self.workbook[self.input_sheet_name]

    def get_models(self) -> dict[str: int]:
        sheet = self.input_sheet

        models_rows = map(lambda row: (row[0], 0 if row[1] is None else row[1]),
                          filter(lambda row: row[0] is not None,
                                 sheet.iter_rows(min_row=self.models_start_cell[0],
                                                 min_col=self.models_start_cell[1],
                                                 max_row=self.models_end_cell[0],
                                                 max_col=self.models_end_cell[1],
                                                 values_only=True)
                                 )
                          )

        return dict(models_rows)

    def create_temp_model_file(self, model: Model, market: Optional[str] = None) -> None:
        temp_workbook = openpyxl.Workbook()
        sheet: Worksheet = temp_workbook.worksheets[0]
        sheet.title = model.name
        sheet.cell(1, 1).value = "Дата"
        sheet.cell(1, 1).hyperlink = "test"
        sheet.sheet_format.baseColWidth = 17

        if market:
            sheet.cell(1, 2, "Цена")
            sheet.title = f"{model.name} ({market})"
            self._load_market_prices(sheet, model.get_history(), market)
        else:
            self._load_markets_prices(sheet, model.get_history())

        filename = f"{self.temp_directory}\\{uuid4()}.xlsx"
        temp_workbook.save(filename)
        os.system(filename)

    @staticmethod
    def _load_market_prices(sheet: Worksheet, history: ModelHistory, market: str) -> None:
        points = history.points
        sheet["A1"].font = Font(bold=True)

        # First point
        sheet.cell(2, 1, points[0].date)
        sheet.cell(2, 2, points[0].prices.get(market))
        sheet[f"A2"].font = Font(bold=True)

        row_index = 3
        for points_index in range(1, len(points)):
            previous_point, current_point = points[points_index - 1], points[points_index]
            previous_price, current_price = previous_point.prices.get(market), current_point.prices.get(market)

            if not current_price or not previous_price:
                continue

            if current_price != previous_price:
                sheet.cell(row_index, 2, current_price)
                sheet.cell(row_index, 1, current_point.date)
                sheet[f"A{row_index}"].font = Font(bold=True)
                row_index += 1

    @staticmethod
    def _load_markets_prices(sheet: Worksheet, history: ModelHistory) -> None:
        # Markets
        for markets_index, market in enumerate(db.markets):
            sheet.cell(1, markets_index + 2, market)

        points = history.points
        sheet["A1"].font = Font(bold=True)

        # First point
        sheet.cell(2, 1, points[0].date)
        sheet[f"A2"].font = Font(bold=True)
        for price_index, price in enumerate(points[0].prices.values()):
            sheet.cell(2, price_index + 2, price)

        row_index = 3
        for points_index in range(1, len(points)):
            previous_point, current_point = points[points_index - 1], points[points_index]
            added = False

            for markets_index, (market, current_price) in enumerate(current_point.prices.items()):
                previous_price = previous_point.prices.get(market)

                if not current_price or not previous_price:
                    continue

                if current_price != previous_price:
                    sheet.cell(row_index, markets_index + 2, current_price)
                    added = True

            if added:
                sheet.cell(row_index, 1, current_point.date)
                sheet[f"A{row_index}"].font = Font(bold=True)
                row_index += 1

    def edit_model_market_cell(self, model_name: str, market_name: str, value: Union[int, str, float]) -> None:
        market_column_index = self.get_market_column_index(market_name)
        model_row_index = self.get_model_row_index(model_name)

        sheet: Worksheet = self.workbook[self.input_sheet_name]
        sheet.cell(model_row_index, market_column_index, value)

    def get_market_column_index(self, market_name: str) -> int:
        sheet = self.input_sheet
        markets_ids_row = list(map(lambda cell: cell.value, list(sheet.rows)[self.markets_names_row - 1]))
        return markets_ids_row.index(market_name) + 1

    def get_model_row_index(self, model_name: str) -> int:
        sheet = self.input_sheet
        models_names_column = list(map(lambda cell: cell.value, list(sheet.columns)[self.models_names_column - 1]))
        return models_names_column.index(model_name) + 1
